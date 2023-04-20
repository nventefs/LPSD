import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import openpyxl
import os

rows        = []
row         = []
filenames           = []
timestamps          = []

def excel_read(datafile):
    wb = openpyxl.load_workbook(datafile, data_only = True) 
    sheet = wb.active 
    maxrow = sheet.max_row
    return [wb, sheet, maxrow]

def row_count(datafile): # count total rows in csv file
    data = pd.read_excel(datafile) 
    row_count = sum(1 for row in data)
    return(row_count)

def test_case(i):
    xlname = xlsx_name(i)
    directory = 'Test Case - ' + str(i)    
    return [xlname, directory]

def xlsx_name(i):
    directory = 'Test Case - '+ str(i)

    global filenames
    global timestamps

    for filename in os.listdir(directory):
        if(filename.endswith('.xlsx')):
            f = os.path.join(directory, filename)
            filenames.append(f)
            d = os.path.getmtime(f)
            timestamps.append(str(d))

    index = 0
    for j in range(len(filenames)-1):
        print("Checking file: {} with timestamp: {}".format(filenames[j],timestamps[j]))       
        if timestamps[j] > timestamps[index]:
            index = j
            print("Most recent file is: {} with timestamp: {}".format(filenames[index],timestamps[index]))

    xl_name = filenames[index]
    xl_name = xl_name.replace('~','')
    xl_name = xl_name.replace('$','')
    print("Read data from {}".format(xl_name))
    return(xl_name)


# DEFINE TEST CASE TO EVALUATE
[xl_name, filename] = test_case(7)  

[wb, sheet, maxrow] = excel_read(xl_name)

# REDUCTIVE

sheet = wb['Reductive']
maxrow = sheet.max_row
print(xl_name)
print(maxrow)
truecount = 0

for k in range(2,maxrow + 1):
    cell = sheet.cell(row = k, column = 12)
    if((str(cell.value) == True) or (str(cell.value) == 'True')):
        truecount = truecount + 1

for k in range(2,maxrow + 1):
    cell = sheet.cell(row = k, column = 17)
    if((str(cell.value) == True) or (str(cell.value) == 'True')):
        truecount = truecount + 1


#filename = (xl_name[14:17] + '_Completion_Chart.png')  #stores chart locally
filename = (xl_name[0:17] + '_Reductive_Chart.png') #stores chart in proper folder
print('Reductive : {} // {}'.format(truecount, maxrow - 2))
y = np.array([truecount, maxrow - 2 - truecount])
pielabels = ['Accurate', 'Inaccurate']
plt.pie(y, labels = pielabels)
plt.title(xl_name[14:17] + 'Reductive')
plt.savefig(filename)
plt.close()

# MULTIPLICATIVE

filename = (xl_name[0:17] + '_Multiplicative_Chart.png')
sheet = wb['Multiplicative']
maxrow = sheet.max_row
truecount = 0

for k in range(2,maxrow + 1):
    cell = sheet.cell(row = k, column = 20)
    if((str(cell.value) == True) or (str(cell.value) == 'True') or (cell.value == True) or (cell.value == 1)):
        truecount = truecount + 1

print('Multiplicative : {} // {}'.format(truecount, maxrow - 2))
y = np.array([truecount, maxrow - 2 - truecount])
pielabels = ['Accurate', 'Inaccurate']
plt.pie(y, labels = pielabels)
plt.title(xl_name[14:17] + 'Multiplicative')
plt.savefig(filename)
plt.close()
