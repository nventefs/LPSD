import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import openpyxl
import os

rows        = []
row         = []


def excel_read(datafile):
    wb = openpyxl.load_workbook(datafile) 
    sheet = wb.active 
    maxrow = sheet.max_row
    return [wb, sheet, maxrow]

def row_count(datafile): # count total rows in csv file
    data = pd.read_excel(datafile) 
    row_count = sum(1 for row in data)
    return(row_count)

def test_case(i):
    xlname = xlsx_name(i)
    directory = 'Test Case - ' + str(i)
    return [xlname, directory]

def xlsx_name(i):
    directory = 'Test Case - '+ str(i)
    for filename in os.listdir(directory):
        f = os.path.join(directory, filename)
        # checking if it is a file
        if (os.path.isfile(f) and f[-5:] == '.xlsx' and f[14:16] == 'TC'):
            return(f)
            #return(f[14:])

# DEFINE TEST CASE TO EVALUATE
[xl_name, filename] = test_case(7)  


[wb, sheet, maxrow] = excel_read(xl_name)
truecount = 0

for k in range(1,maxrow + 1):
    cell = sheet.cell(row = k, column = 12)
    if(str(cell.value) == 'True'):
        truecount = truecount + 1

for k in range(1,maxrow + 1):
    cell = sheet.cell(row = k, column = 17)
    if(str(cell.value) == 'True'):
        truecount = truecount + 1


#filename = (xl_name[14:17] + '_Completion_Chart.png')  #stores chart locally
filename = (xl_name[0:17] + '_Completion_Chart.png') #stores chart in proper folder

y = np.array([truecount, maxrow - 2 - truecount])
pielabels = ['Accurate', 'Inaccurate']
plt.pie(y, labels = pielabels)
plt.title(xl_name[14:17])
plt.savefig(filename)
plt.close()